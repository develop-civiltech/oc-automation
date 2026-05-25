#!/usr/bin/env python3
"""
generarOC.py v6 — Para plantilla limpia sin vínculos externos.
Usa openpyxl directamente ya que la plantilla no tiene externalLinks.
Preserva fórmulas simples (K33, K34, K35) para que Excel las calcule.
"""
import sys, json, os, shutil
from datetime import datetime
from pathlib import Path

def main():
    try:
        raw = sys.stdin.buffer.read()
        entrada = json.loads(raw.decode('utf-8'))
        plantilla = Path(entrada['plantilla'])
        salida    = Path(entrada['salida'])
        datos     = entrada['datos']

        if not plantilla.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {plantilla}")

        salida.parent.mkdir(parents=True, exist_ok=True)

        from openpyxl import load_workbook

        shutil.copy2(str(plantilla), str(salida))
        wb = load_workbook(str(salida))
        ws = wb['OC']

        def w(celda, valor):
            """Escribe valor respetando celdas fusionadas."""
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_str, row = coordinate_from_string(celda)
            col = column_index_from_string(col_str)
            for rango in ws.merged_cells.ranges:
                if rango.min_row <= row <= rango.max_row and rango.min_col <= col <= rango.max_col:
                    ws.cell(rango.min_row, rango.min_col).value = valor
                    return
            ws[celda].value = valor

        PCT_IVA = 0.19

        # Cabecera
        w('J7', datos.get('numeroOC', ''))
        w('J8', datos.get('fecha', ''))
        w('J9', datos.get('proyecto', ''))

        # Proveedor
        prov = dados = datos.get('proveedor') or {}
        w('B12', prov.get('nombre', 'PENDIENTE - COTIZACION REQUERIDA'))
        w('E12', prov.get('nit', ''))
        w('H12', 'Asesor')
        w('K12', prov.get('municipio', ''))
        w('B13', prov.get('telefono', ''))
        w('E13', prov.get('telefono', ''))
        w('H13', prov.get('correo', ''))

        # Ítems
        items = datos.get('items', [])
        total_sin_iva = 0
        total_iva     = 0

        for i in range(16):
            fila = 16 + i
            if i < len(items):
                item = items[i]
                try:    precio = float(item.get('precio') or 0)
                except: precio = 0
                try:    cant = float(item.get('cantidad') or 1)
                except: cant = 1

                ws[f'B{fila}'].value = item.get('insumo', '')
                ws[f'F{fila}'].value = item.get('unidad', 'UND')
                ws[f'G{fila}'].value = cant

                if precio > 0:
                    sin_iva  = round(precio / (1 + PCT_IVA), 2)
                    subtotal = round(precio * cant, 2)
                    iva_item = round(subtotal - sin_iva * cant, 2)
                    ws[f'H{fila}'].value = round(precio, 2)
                    ws[f'I{fila}'].value = 0
                    ws[f'J{fila}'].value = PCT_IVA
                    ws[f'K{fila}'].value = subtotal
                    total_sin_iva += sin_iva * cant
                    total_iva     += iva_item
                else:
                    ws[f'H{fila}'].value = None
                    ws[f'I{fila}'].value = 0
                    ws[f'J{fila}'].value = 0
                    ws[f'K{fila}'].value = 0
            else:
                ws[f'B{fila}'].value = None
                ws[f'F{fila}'].value = None
                ws[f'G{fila}'].value = None
                ws[f'H{fila}'].value = None
                ws[f'I{fila}'].value = 0
                ws[f'J{fila}'].value = 0
                ws[f'K{fila}'].value = 0

        # Totales — escribir valores directos (las fórmulas en la plantilla los calcularán)
        ws['K33'].value = round(total_sin_iva, 2)
        ws['K34'].value = round(total_iva, 2)
        ws['K35'].value = round(total_sin_iva + total_iva, 2)

        # Firma y fecha
        ws['H48'].value = 'ING. BRAYAN ALEXANDER OSPINA VASQUEZ'
        ws['H49'].value = 'COORDINADOR DE PROYECTOS'
        ws['K51'].value = datetime.now().strftime('%d/%m/%Y')

        wb.save(str(salida))

        if not salida.exists():
            raise FileNotFoundError(f"Archivo no creado: {salida}")

        print(json.dumps({'ok': True, 'ruta': str(salida)}, ensure_ascii=False))

    except Exception as e:
        import traceback
        print(json.dumps({
            'ok': False,
            'error': str(e),
            'detalle': traceback.format_exc()
        }, ensure_ascii=False))
        sys.exit(1)

main()